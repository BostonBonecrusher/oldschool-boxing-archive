"""
mine_stories.py
===============
EXHAUSTIVE story miner -- now works against EITHER archive collection
(BoxingChunk or HistoricalWarfareChunk) via --collection.

The problem with a pure search-based approach (mine_story_content.py /
mine_gladiator_mentions.py): firing a handful of hybrid SEARCH queries and
keeping the top hits always rewards the same strongest passages, so you
keep getting the same famous stories back. That is re-milking the same
dozen cows.

This script does the opposite. It does NOT search. It pages through EVERY
chunk in the chosen collection exactly once (cursor sweep), so by
construction it visits all material -- the obscure single-book anecdotes
as well as the famous ones. Each chunk is read by a cheap LLM that asks
one question:

    "Is there a vivid, specific, surprising STORY here -- the kind that
     makes a good Instagram reel? If so, extract it."

Accepted stories are logged, then a dedup pass collapses the same story
told across many books into ONE entry (and counts the corroborations,
which is a trust signal). Final output is a sortable Excel story bank with
full citations so every reel claim is traceable to book + author + year +
page.

Every export (--export-only or after a sweep) also refreshes a shared
story_hooks.json -- a curated subset (wow >= STORY_HOOKS_MIN_WOW) that the
live web app reads from for its "Story Ideas" (random), "Browse by name",
and "Filter by era" features. No extra command needed; it's written
automatically alongside the .xlsx, and each collection only ever touches
its own slice of that file.

----------------------------------------------------------------------
COLLECTIONS
----------------------------------------------------------------------
--collection boxing    (default) -- BoxingChunk. Has Phase-2 tags already
                        run (contains_fight_account etc.), so the cheap
                        "story" pre-filter works and is the default.
--collection warfare   -- HistoricalWarfareChunk. Phase-2 AI-tagging has
                        NOT been run on this collection yet, so there is
                        no cheap pre-filter available -- this mode
                        defaults to --filter all (every chunk gets fed to
                        the LLM, which costs more per chunk swept).
--collection both      -- runs boxing, then warfare, one after another, in
                        this SAME command (one process, two passes).
                        --limit/--filter/--reset apply identically to each
                        pass (e.g. --limit 3000 --collection both sweeps up
                        to 3000 chunks of boxing AND up to 3000 of warfare).

Each collection writes to its OWN log/raw/xlsx files, so running one never
touches the other's progress:
    boxing  -> mine_stories_log.json / mine_stories_raw.jsonl / story_bank.xlsx
    warfare -> mine_stories_log_warfare.json / mine_stories_raw_warfare.jsonl
               / story_bank_warfare.xlsx

----------------------------------------------------------------------
WORKFLOW
----------------------------------------------------------------------
Proof-of-concept (sample a few thousand chunks, ~minutes):
    python mine_stories.py --collection boxing  --limit 3000
    python mine_stories.py --collection warfare --limit 3000
    python mine_stories.py --collection both    --limit 3000   # both, one command

Full run (whole collection -- long; safe to stop/restart, it resumes):
    python mine_stories.py --collection warfare --full
    python mine_stories.py --collection both    --full   # both, back to back

Re-export the Excel bank from what's already been mined (no API calls):
    python mine_stories.py --collection boxing  --export-only
    python mine_stories.py --collection warfare --export-only

Start over from scratch (wipes that collection's progress log + raw finds):
    python mine_stories.py --collection warfare --reset --limit 3000

Useful flags:
    --filter story   only feed chunks tagged as story-bearing (cheap;
                      boxing default -- ignored/forced to "all" for
                      warfare since it has no tags yet)
    --filter all     feed every chunk regardless of tags (most thorough,
                      pricier; warfare default)
    --min-score 6    only export stories scored >= this (default 6)
    --batch 8        chunks per LLM call (default 8)
    --model gpt-4o-mini   extraction model (default)

Outputs (written next to this script, namespaced per collection as above).
----------------------------------------------------------------------
"""

import os
import sys
import json
import time
import argparse
from datetime import datetime

from dotenv import load_dotenv
import weaviate
from weaviate.classes.init import Auth
from openai import OpenAI

load_dotenv()

WEAVIATE_URL     = os.getenv("WEAVIATE_URL")
WEAVIATE_API_KEY = os.getenv("WEAVIATE_API_KEY")
OPENAI_API_KEY   = os.getenv("OPENAI_API_KEY")

HERE = os.path.dirname(os.path.abspath(__file__))

EMBED_MODEL = "text-embedding-3-small"   # same model your archive already uses
DEDUP_THRESHOLD = 0.86                    # cosine >= this => "same story"

# Shared data file the live web app reads from for its "Story Ideas" /
# "Browse by name" features. One file, keyed by collection, so an export
# of one collection never clobbers the other's already-curated stories.
STORY_HOOKS_FILE = os.path.join(HERE, "story_hooks.json")
STORY_HOOKS_MIN_WOW = 7    # only the best stories get surfaced to visitors


# ─────────────────────────────────────────────────────────────────────
# EXTRACTION PROMPTS -- one per collection, same shape, different voice
# and example stories so the model knows what "good" looks like in that
# domain.
# ─────────────────────────────────────────────────────────────────────
SYSTEM_PROMPT_BOXING = """You are a boxing historian mining old books for SHORT, \
PUNCHY, SURPRISING true anecdotes to turn into Instagram reels.

You will receive several numbered passages from old boxing books. For EACH
passage, decide whether it contains a genuinely eye-catching STORY or fact --
the kind that makes someone stop scrolling. Good examples of what to look for:
- "When Jack Dempsey was a hobo he was robbed, then tracked the robbers down."
- "Jeffries said Fitzsimmons hit like a man who weighed 300 pounds."
- "Jack Johnson ran 15 miles in wool on an empty stomach before fights."
Vivid, specific, unusual, human. Concrete details, named people, surprising
turns, wild training methods, feuds, near-death moments, quotes, oddities.

REJECT passages that are: generic overviews, dry record-keeping, lists of
results, tables of contents, indexes, OCR garble, or anything you could not
build a single compelling caption around. Be selective -- a no is fine.

Return STRICT JSON, an array with one object PER PASSAGE THAT CONTAINS A
STORY (skip the rest entirely). Each object:
{
  "n": <passage number>,
  "hook": "<one punchy sentence, <=160 chars, the reel headline>",
  "detail": "<2-4 sentences of the story, faithful to the passage>",
  "people": ["<fighter or person names involved>"],
  "category": "<one of: training | fight | biography | feud | quote | oddity | death | money | crime>",
  "wow": <integer 1-10, how eye-catching/scroll-stopping this is>,
  "specificity": <integer 1-10, how concrete & verifiable vs vague>
}
If a passage has NO worthwhile story, do not include an object for it.
Output ONLY the JSON array, nothing else."""

SYSTEM_PROMPT_WARFARE = """You are a military historian mining old books about \
gladiators, warriors, and armies (any era) for SHORT, PUNCHY, SURPRISING true \
anecdotes to turn into Instagram reels.

You will receive several numbered passages from old history books. For EACH
passage, decide whether it contains a genuinely eye-catching STORY or fact --
the kind that makes someone stop scrolling. Good examples of what to look for:
- "Spartacus, once a gladiator, led a slave army that beat Roman legions in
   open battle before his final defeat."
- "When a legion lost its eagle standard, soldiers volunteered for suicide
   missions to recover it -- losing the eagle was a disgrace worse than death."
- "A Roman general decimated his own retreating troops -- executing one in
   ten men, chosen by lottery, by their own comrades' hands."
Vivid, specific, unusual, human. Concrete details, named people/units,
surprising turns, brutal discipline, famous duels, training methods, oaths
and rituals, mutinies, sieges, oddities.

REJECT passages that are: generic overviews, dry record-keeping, lists of
troop counts/dates, tables of contents, indexes, OCR garble, or anything you
could not build a single compelling caption around. Be selective -- a no is
fine.

Return STRICT JSON, an array with one object PER PASSAGE THAT CONTAINS A
STORY (skip the rest entirely). Each object:
{
  "n": <passage number>,
  "hook": "<one punchy sentence, <=160 chars, the reel headline>",
  "detail": "<2-4 sentences of the story, faithful to the passage>",
  "people": ["<warrior, commander, gladiator, or unit/legion names involved>"],
  "category": "<one of: training | combat | biography | feud | quote | oddity | death | mutiny | siege | ritual>",
  "wow": <integer 1-10, how eye-catching/scroll-stopping this is>,
  "specificity": <integer 1-10, how concrete & verifiable vs vague>
}
If a passage has NO worthwhile story, do not include an object for it.
Output ONLY the JSON array, nothing else."""


# ─────────────────────────────────────────────────────────────────────
# COLLECTION CONFIG -- everything that differs between archives lives
# here. Add a new collection by adding a new entry to this dict.
# ─────────────────────────────────────────────────────────────────────
COLLECTIONS = {
    "boxing": {
        "key": "boxing",
        "weaviate_name": "BoxingChunk",
        "return_props": [
            "content", "source_file", "title", "author", "year_published",
            "is_primary_source", "era", "page_number", "chunk_index",
            "subject_fighter", "fighters_mentioned",
            "contains_fight_account", "contains_biographical_info",
            "quotes_present", "controversy_present", "contains_training_methods",
        ],
        # A chunk is "story-bearing" if any of these Phase-2 flags is true.
        "story_flags": [
            "contains_fight_account", "contains_biographical_info",
            "quotes_present", "controversy_present", "contains_training_methods",
        ],
        "system_prompt": SYSTEM_PROMPT_BOXING,
        "log_file": os.path.join(HERE, "mine_stories_log.json"),
        "raw_file": os.path.join(HERE, "mine_stories_raw.jsonl"),
        "xlsx_file": os.path.join(HERE, "story_bank.xlsx"),
        "default_filter": "story",
    },
    "warfare": {
        "key": "warfare",
        "weaviate_name": "HistoricalWarfareChunk",
        "return_props": [
            "content", "source_file", "title", "author", "year_published",
            "is_primary_source", "era", "civilization_or_nation",
            "conflict_focus", "page_number", "chunk_index",
            "subject_commander", "discipline",
        ],
        # Phase-2 AI-tagging has NOT been run on this collection yet, so
        # there is no cheap boolean pre-filter available. Left empty on
        # purpose -- is_story_bearing() always returns False for an empty
        # list, which is exactly right: with no tags, "story" mode would
        # silently feed nothing, so the default filter below is "all".
        "story_flags": [],
        "system_prompt": SYSTEM_PROMPT_WARFARE,
        "log_file": os.path.join(HERE, "mine_stories_log_warfare.json"),
        "raw_file": os.path.join(HERE, "mine_stories_raw_warfare.jsonl"),
        "xlsx_file": os.path.join(HERE, "story_bank_warfare.xlsx"),
        "default_filter": "all",
    },
}


# ─────────────────────────────────────────────────────────────────────
# PROGRESS LOG (per-collection files passed in via cfg)
# ─────────────────────────────────────────────────────────────────────
def load_seen(log_file):
    if os.path.exists(log_file):
        with open(log_file, "r", encoding="utf-8") as f:
            return set(json.load(f).get("seen_uuids", []))
    return set()


def save_seen(log_file, seen):
    with open(log_file, "w", encoding="utf-8") as f:
        json.dump({"seen_uuids": sorted(seen),
                   "updated": datetime.now().isoformat()}, f)


def append_raw(raw_file, stories):
    with open(raw_file, "a", encoding="utf-8") as f:
        for s in stories:
            f.write(json.dumps(s, ensure_ascii=False) + "\n")


# ─────────────────────────────────────────────────────────────────────
# LLM EXTRACTION
# ─────────────────────────────────────────────────────────────────────
def extract_from_batch(oai, model, system_prompt, batch):
    """batch = list of (uuid, props). Returns list of story dicts."""
    passages = []
    for i, (_uid, p) in enumerate(batch, 1):
        text = (p.get("content") or "").replace("\n", " ").strip()
        passages.append(f"[Passage {i}]\n{text[:1800]}")
    user_msg = "\n\n".join(passages)

    try:
        resp = oai.chat.completions.create(
            model=model,
            temperature=0.2,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_msg},
            ],
            response_format={"type": "json_object"},
        )
        raw = resp.choices[0].message.content.strip()
    except Exception as e:
        print(f"   ! LLM error on batch: {str(e)[:120]}")
        return []

    # The model may wrap the array in an object; handle both.
    try:
        data = json.loads(raw)
        if isinstance(data, dict):
            arr = next((v for v in data.values() if isinstance(v, list)), [])
        else:
            arr = data
    except Exception:
        return []

    out = []
    for item in arr:
        if not isinstance(item, dict):
            continue
        n = item.get("n")
        if not isinstance(n, int) or n < 1 or n > len(batch):
            continue
        uid, p = batch[n - 1]
        out.append({
            "uuid": uid,
            "hook": (item.get("hook") or "").strip(),
            "detail": (item.get("detail") or "").strip(),
            "people": item.get("people") or [],
            "category": (item.get("category") or "").strip(),
            "wow": int(item.get("wow") or 0),
            "specificity": int(item.get("specificity") or 0),
            # citation carried straight from the chunk:
            "title": p.get("title") or p.get("source_file", "Unknown"),
            "author": p.get("author", ""),
            "year": p.get("year_published", 0),
            "source_file": p.get("source_file", ""),
            "page": p.get("page_number", 0),
            "chunk_index": p.get("chunk_index", 0),
            "is_primary_source": bool(p.get("is_primary_source", False)),
            "era": p.get("era", ""),
        })
    return [s for s in out if s["hook"]]


def is_story_bearing(props, story_flags):
    if not story_flags:
        return False
    return any(bool(props.get(flag)) for flag in story_flags)


# ─────────────────────────────────────────────────────────────────────
# SWEEP  (exhaustive cursor pass over the whole collection)
# ─────────────────────────────────────────────────────────────────────
def sweep(args, cfg):
    oai = OpenAI(api_key=OPENAI_API_KEY)
    print(f"Connecting to Weaviate Cloud ... (collection: {cfg['weaviate_name']})")
    client = weaviate.connect_to_weaviate_cloud(
        cluster_url=WEAVIATE_URL,
        auth_credentials=Auth.api_key(WEAVIATE_API_KEY),
        headers={"X-OpenAI-Api-Key": OPENAI_API_KEY},
    )
    try:
        collection = client.collections.get(cfg["weaviate_name"])
        total = collection.aggregate.over_all(total_count=True).total_count
        print(f"Connected. {cfg['weaviate_name']} holds {total:,} chunks.")

        seen = load_seen(cfg["log_file"])
        print(f"Already processed in a previous run: {len(seen):,} chunks.")
        print(f"Filter mode: {args.filter}  |  cap this run: "
              f"{'none (full)' if args.full else args.limit}\n")

        batch = []
        processed_this_run = 0
        fed = 0           # chunks actually sent to the LLM
        found = 0         # stories accepted
        t0 = time.time()

        def flush():
            nonlocal found
            if not batch:
                return
            stories = extract_from_batch(oai, args.model, cfg["system_prompt"], batch)
            if stories:
                append_raw(cfg["raw_file"], stories)
                found += len(stories)
                for s in stories:
                    tag = "PRIMARY" if s["is_primary_source"] else ""
                    print(f"   + [{s['wow']}/10] {s['hook'][:90]}  "
                          f"({s['title'][:40]} {tag})")
            for uid, _ in batch:
                seen.add(uid)
            save_seen(cfg["log_file"], seen)
            batch.clear()

        # iterator() is a cursor over EVERY object -- the exhaustive sweep.
        for obj in collection.iterator(
            include_vector=False, return_properties=cfg["return_props"]
        ):
            uid = str(obj.uuid)
            if uid in seen:
                continue
            props = obj.properties

            if args.filter == "story" and not is_story_bearing(props, cfg["story_flags"]):
                seen.add(uid)              # mark seen so we never re-check it
                processed_this_run += 1
                continue

            batch.append((uid, props))
            fed += 1
            processed_this_run += 1

            if len(batch) >= args.batch:
                flush()
                rate = processed_this_run / max(time.time() - t0, 1)
                print(f"   ... {processed_this_run:,} swept this run | "
                      f"{fed:,} fed to LLM | {found:,} stories | "
                      f"{rate:.0f} chunks/s", end="\r")

            if not args.full and processed_this_run >= args.limit:
                break

        flush()
        print("\n\n" + "=" * 64)
        print(f"Sweep done. Swept {processed_this_run:,} chunks this run, "
              f"fed {fed:,} to the LLM, accepted {found:,} stories.")
        print(f"Raw finds appended to {os.path.basename(cfg['raw_file'])}")
        print("=" * 64 + "\n")
    finally:
        client.close()


# ─────────────────────────────────────────────────────────────────────
# DEDUP + EXPORT
# ─────────────────────────────────────────────────────────────────────
def load_raw(raw_file):
    if not os.path.exists(raw_file):
        return []
    rows = []
    with open(raw_file, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    rows.append(json.loads(line))
                except Exception:
                    pass
    return rows


def update_story_hooks(coll_key, records):
    """
    Refresh this collection's slice of the shared story_hooks.json -- the
    data file the live web app reads for its "Story Ideas" (random) and
    "Browse by name" features. Only stories scored >= STORY_HOOKS_MIN_WOW
    make the cut, since this file is visitor-facing (a curated subset, not
    the whole bank). The other collection's entries in the file are left
    untouched, so exporting boxing never clobbers warfare's hooks and
    vice versa.
    """
    data = {}
    if os.path.exists(STORY_HOOKS_FILE):
        try:
            with open(STORY_HOOKS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            data = {}

    picks = []
    for r in records:
        if r["wow"] < STORY_HOOKS_MIN_WOW:
            continue
        people = [p for p in r["people"].split(", ") if p]
        picks.append({
            "hook": r["hook"],
            "category": r["category"],
            "people": people,
            "wow": r["wow"],
            "era": r.get("era", ""),
        })

    data[coll_key] = picks
    with open(STORY_HOOKS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"-> {STORY_HOOKS_FILE}  ({len(picks):,} stories at wow>={STORY_HOOKS_MIN_WOW} "
          f"for '{coll_key}', for the live app's Story Ideas button)")


def dedup_and_export(args, cfg):
    import numpy as np
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    rows = load_raw(cfg["raw_file"])
    if not rows:
        print(f"No raw stories found yet for {cfg['weaviate_name']}. Run a sweep first.")
        return

    print(f"Loaded {len(rows):,} raw stories. Embedding hooks for dedup ...")
    oai = OpenAI(api_key=OPENAI_API_KEY)

    hooks = [r["hook"] for r in rows]
    vecs = []
    B = 256
    for i in range(0, len(hooks), B):
        chunk = hooks[i:i + B]
        resp = oai.embeddings.create(model=EMBED_MODEL, input=chunk)
        vecs.extend([d.embedding for d in resp.data])
    V = np.array(vecs, dtype=np.float32)
    V /= (np.linalg.norm(V, axis=1, keepdims=True) + 1e-9)

    # Greedy clustering: each story joins the first cluster it's close to.
    order = sorted(range(len(rows)), key=lambda i: rows[i]["wow"], reverse=True)
    cluster_of = [-1] * len(rows)
    centroids = []          # list of (cluster_id, vector)
    clusters = []           # list of list[idx]
    for i in order:
        v = V[i]
        best_c, best_sim = -1, -1.0
        for cid, cvec in centroids:
            sim = float(np.dot(v, cvec))
            if sim > best_sim:
                best_sim, best_c = sim, cid
        if best_sim >= DEDUP_THRESHOLD:
            cluster_of[i] = best_c
            clusters[best_c].append(i)
        else:
            cid = len(clusters)
            clusters.append([i])
            centroids.append((cid, v))
            cluster_of[i] = cid

    print(f"Collapsed into {len(clusters):,} distinct stories "
          f"(from {len(rows):,} raw finds).")

    # Build one row per cluster, choosing the highest-wow representative.
    records = []
    for cid, idxs in enumerate(clusters):
        idxs_sorted = sorted(idxs, key=lambda i: rows[i]["wow"], reverse=True)
        rep = rows[idxs_sorted[0]]
        if rep["wow"] < args.min_score:
            continue
        cites, people = [], []
        for i in idxs:
            r = rows[i]
            c = r["title"]
            if r.get("author"):
                c += f" — {r['author']}"
            if r.get("year"):
                c += f" ({r['year']})"
            if r.get("page"):
                c += f", p.{r['page']}"
            if c not in cites:
                cites.append(c)
            for person in r.get("people", []):
                if person and person not in people:
                    people.append(person)
        records.append({
            "hook": rep["hook"],
            "detail": rep["detail"],
            "category": rep["category"],
            "people": ", ".join(people[:8]),
            "wow": rep["wow"],
            "specificity": rep["specificity"],
            "corroborations": len(idxs),
            "primary": "yes" if any(rows[i]["is_primary_source"] for i in idxs) else "",
            "best_source": cites[0] if cites else "",
            "all_sources": " | ".join(cites[:10]),
            "era": rep.get("era", ""),
        })

    records.sort(key=lambda r: (r["wow"], r["corroborations"]), reverse=True)
    print(f"{len(records):,} stories scored >= {args.min_score} -> writing Excel.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Story Bank"
    headers = [
        ("hook", 60), ("detail", 70), ("category", 12), ("people", 26),
        ("wow", 6), ("specificity", 7), ("corroborations", 8),
        ("primary", 8), ("era", 16), ("best_source", 45), ("all_sources", 80),
    ]
    head_fill = PatternFill("solid", fgColor="1F2A44")
    head_font = Font(bold=True, color="FFFFFF")
    for c, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = width
    for r, rec in enumerate(records, 2):
        for c, (name, _w) in enumerate(headers, 1):
            val = rec.get(name, "")
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(wrap_text=name in ("hook", "detail", "all_sources"),
                                       vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(records)+1}"
    wb.save(cfg["xlsx_file"])
    print(f"\nDone -> {cfg['xlsx_file']}")

    update_story_hooks(cfg["key"], records)


# ─────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="Exhaustive story miner (boxing or warfare archive)")
    ap.add_argument("--collection", choices=list(COLLECTIONS.keys()) + ["both"], default="boxing",
                    help="Which archive to mine: 'boxing' (BoxingChunk), "
                         "'warfare' (HistoricalWarfareChunk), or 'both' (runs "
                         "boxing then warfare, one after another, in this same "
                         "command). Default: boxing.")
    ap.add_argument("--limit", type=int, default=3000,
                    help="max chunks to sweep this run (ignored with --full)")
    ap.add_argument("--full", action="store_true",
                    help="sweep the entire collection (long; resumable)")
    ap.add_argument("--filter", choices=["story", "all"], default=None,
                    help="'story' = only tagged story-bearing chunks (cheap, boxing "
                         "only -- warfare has no tags yet); 'all' = every chunk "
                         "(most thorough). If omitted, uses the collection's default "
                         "(boxing='story', warfare='all').")
    ap.add_argument("--batch", type=int, default=8,
                    help="chunks per LLM call")
    ap.add_argument("--model", default="gpt-4o-mini",
                    help="extraction model")
    ap.add_argument("--min-score", type=int, default=6,
                    help="only export stories with wow >= this")
    ap.add_argument("--export-only", action="store_true",
                    help="skip the sweep; just dedup+export what's already mined")
    ap.add_argument("--reset", action="store_true",
                    help="wipe this collection's progress log and raw finds before running")
    args = ap.parse_args()

    if not (WEAVIATE_URL and WEAVIATE_API_KEY and OPENAI_API_KEY):
        sys.exit("Missing WEAVIATE_URL / WEAVIATE_API_KEY / OPENAI_API_KEY in .env")

    targets = list(COLLECTIONS.keys()) if args.collection == "both" else [args.collection]
    for i, coll_key in enumerate(targets, 1):
        if len(targets) > 1:
            print("\n" + "#" * 72)
            print(f"# PASS {i}/{len(targets)}: {coll_key} "
                  f"({COLLECTIONS[coll_key]['weaviate_name']})")
            print("#" * 72 + "\n")
        run_one(args, coll_key)


def run_one(args, coll_key):
    """Resolve one collection's config and run sweep+export for it. Called
    directly for --collection boxing/warfare, and once per item in the loop
    above for --collection both."""
    cfg = COLLECTIONS[coll_key]

    filt = args.filter
    if filt is None:
        filt = cfg["default_filter"]
    if filt == "story" and not cfg["story_flags"]:
        print(f"NOTE: '{coll_key}' has no story_flags defined (Phase-2 tags "
              f"not run yet), so --filter story would feed nothing. Forcing --filter all.\n")
        filt = "all"

    if args.reset:
        for f in (cfg["log_file"], cfg["raw_file"]):
            if os.path.exists(f):
                os.remove(f)
        print(f"Reset: cleared progress log and raw finds for '{coll_key}'.\n")

    # sweep()/dedup_and_export() read .filter off the args object -- pass a
    # per-pass copy with the resolved value so --both can't leak boxing's
    # resolved filter into the warfare pass (or vice versa).
    run_args = argparse.Namespace(**vars(args))
    run_args.filter = filt

    if not args.export_only:
        sweep(run_args, cfg)
    dedup_and_export(run_args, cfg)


if __name__ == "__main__":
    main()
